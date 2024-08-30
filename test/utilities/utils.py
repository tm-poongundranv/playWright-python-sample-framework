import logging
import configparser
import csv
from openpyxl import Workbook, load_workbook


class Utils:

    def get_config_value(section, key, filename='config.ini'):
        """
        Read a value from a configuration file.

        :param section: The section in the config file.
        :param key: The key within the section to read the value.
        :param filename: The name of the config file (default is 'config.ini').
        :return: The value associated with the given section and key.
        """
        # config = configparser.ConfigParser()
        # # Ensure the config file path is correct
        # config_path = "/Users/bhuvaneshwarichilla/PycharmProjects/pythonProject/configurations/config.ini"
        # config.read(config_path)
        #
        # if config.has_section(section) and config.has_option(section, key):
        #     return config.get(section, key)
        # else:
        #     raise configparser.NoSectionError(f"Section '{section}' or key '{key}' not found in '{filename}'.")

    @staticmethod
    def setup_logging(name=None):
        logger = logging.getLogger(name)
        if not logger.handlers:
            fhandler = logging.FileHandler(filename='automationlss.log', mode='a')
            formatter = logging.Formatter('%(asctime)s: %(name)s: %(levelname)s: %(message)s',
                                          datefmt='%m/%d/%Y %I:%M:%S %p')
            fhandler.setFormatter(formatter)
            logger.addHandler(fhandler)
            logger.setLevel(logging.INFO)
        return logger

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = []
            for j in range(1, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            datalist.append(row)
        return datalist

    def read_data_from_csv(filename):
        #Create an empty list
        datalist = []
        #Open CSV file
        csvdata = open(filename, "r")
        #Create CSV reader
        reader = csv.reader(csvdata)
        #skip header
        next(reader)
        #Add CSV rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist

